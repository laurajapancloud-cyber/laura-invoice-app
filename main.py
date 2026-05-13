import os
import json
import logging
import secrets
from typing import Annotated, List, Dict, Any, Optional
import base64
from io import BytesIO
import datetime
from zoneinfo import ZoneInfo
import time
import re
import requests
import uuid
from html import escape

# psycopg2 / supabase はオプション (ローカル UI プレビューでは不要)
try:
    import psycopg2
    from psycopg2 import pool
    from psycopg2.extras import RealDictCursor
except ImportError:
    psycopg2 = None
    pool = None
    RealDictCursor = None

try:
    from supabase import create_client, Client as SupabaseClient
except ImportError:
    create_client = None
    SupabaseClient = None

# ==================== Logging Setup ====================
logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO").upper(),
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("laura")

from fastapi import FastAPI, UploadFile, File, Depends, HTTPException, status, Response, Request, Form, BackgroundTasks
from fastapi.staticfiles import StaticFiles
from fastapi.security import APIKeyCookie
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, RedirectResponse, StreamingResponse, Response
from fastapi.concurrency import run_in_threadpool
import zipfile
import urllib.parse
import hashlib
try:
    import google.generativeai as genai
except ImportError:
    genai = None
try:
    from weasyprint import HTML
except Exception as e:
    logger.error("WeasyPrint import failed (PDF generation will be disabled): %s", e)
    HTML = None

from dotenv import load_dotenv

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.utils.units import pixels_to_EMU
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
except ImportError:
    openpyxl = None

try:
    import barcode
    from barcode.writer import ImageWriter
except ImportError:
    barcode = None
from pydantic import BaseModel

def get_jst_now():
    return datetime.datetime.now(ZoneInfo("Asia/Tokyo"))

def to_jst(dt):
    """DBから来たdatetimeをJSTへ正規化"""
    if not dt:
        return get_jst_now()
    if isinstance(dt, str):
        try:
            dt = datetime.datetime.fromisoformat(dt.replace("Z", "+00:00"))
        except Exception:
            return get_jst_now()
    if isinstance(dt, datetime.datetime):
        JST = ZoneInfo("Asia/Tokyo")
        if dt.tzinfo is None:
            return dt.replace(tzinfo=JST)
        return dt.astimezone(JST)
    return get_jst_now()

def h(value) -> str:
    """HTMLエスケープヘルパー"""
    return escape(str(value or ""), quote=True)

# Conditional imports
try:
    from openai import OpenAI, AzureOpenAI
except ImportError:
    OpenAI = None
    AzureOpenAI = None

try:
    from google.oauth2 import service_account
    from google.cloud import vision
    from googleapiclient.discovery import build as gdrive_build
    from googleapiclient.http import MediaInMemoryUpload
except ImportError:
    vision = None
    gdrive_build = None
    MediaInMemoryUpload = None

# Load environment variables
load_dotenv()

APP_USERNAME = os.getenv("APP_USERNAME")
APP_PASSWORD = os.getenv("APP_PASSWORD")
# テストモード（認証情報なしで test/test ログインを許可）。明示的に "true" を指定した時のみ有効。
DEV_MODE = os.getenv("DEV_MODE", "").lower() == "true"
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
GOOGLE_CREDENTIALS_JSON = os.getenv("GOOGLE_CREDENTIALS_JSON")
AZURE_OPENAI_KEY = os.getenv("AZURE_OPENAI_KEY")
AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
GDRIVE_FOLDER_ID = os.getenv("GDRIVE_FOLDER_ID")
GDRIVE_WEBHOOK_URL = os.getenv("GDRIVE_WEBHOOK_URL")

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY")
STORAGE_BUCKET = "invoices"

supabase_client: Optional["SupabaseClient"] = None
if SUPABASE_URL and SUPABASE_SERVICE_KEY and create_client:
    supabase_client = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

DOC_TYPE_PREFIXES = {
    "delivery": "LJ", "return": "LR",
    "prov_delivery": "TLJ", "prov_return": "TLR",
}
DOC_TYPE_LABELS = {
    "delivery": "納品伝票", "return": "返品伝票",
    "prov_delivery": "仮納品", "prov_return": "仮返品",
}
DOC_TYPE_TITLES = {
    "delivery":      {"main": "納 品 書", "detail": "商 品 明 細 表",   "pdf_title": "納品書",   "detail_pdf_title": "商品明細表"},
    "return":        {"main": "返 品 書", "detail": "返 品 明 細 表",   "pdf_title": "返品書",   "detail_pdf_title": "返品明細表"},
    "prov_delivery": {"main": "仮 納 品 書", "detail": "仮納品 明細表",     "pdf_title": "仮納品書", "detail_pdf_title": "仮納品明細表"},
    "prov_return":   {"main": "仮 返 品 書", "detail": "仮返品 明細表",     "pdf_title": "仮返品書", "detail_pdf_title": "仮返品明細表"},
}

def storage_upload(path: str, data: bytes, mime: str) -> str:
    """Supabase Storage にアップロードし、保存パスを返す"""
    if not supabase_client:
        raise Exception("Supabase Storage が未設定です")
    supabase_client.storage.from_(STORAGE_BUCKET).upload(
        path, data, file_options={"content-type": mime, "upsert": "true"}
    )
    return path

def storage_download(path: str) -> bytes:
    """Storage から bytes を取得"""
    if not supabase_client:
        raise Exception("Supabase Storage が未設定です")
    return supabase_client.storage.from_(STORAGE_BUCKET).download(path)

# AI Initialization
gemini_model = None
if genai and GEMINI_API_KEY:
    try:
        genai.configure(api_key=GEMINI_API_KEY)
        gemini_model = genai.GenerativeModel("gemini-2.5-flash")
    except Exception as e:
        logger.warning("Gemini init failed: %s", e)

openai_client = None
if OPENAI_API_KEY and OpenAI:
    openai_client = OpenAI(api_key=OPENAI_API_KEY)

azure_client = None
if AZURE_OPENAI_KEY and AZURE_OPENAI_ENDPOINT and AzureOpenAI:
    # URLの末尾が /openai の場合は自動調整 (AzureOpenAIが内部で付与するため)
    base_endpoint = AZURE_OPENAI_ENDPOINT.rstrip('/')
    if base_endpoint.endswith('/openai'):
        base_endpoint = base_endpoint[:-7]
        
    azure_client = AzureOpenAI(
        api_key=AZURE_OPENAI_KEY,
        api_version="2024-02-01",
        azure_endpoint=base_endpoint,
        default_headers={"Ocp-Apim-Subscription-Key": AZURE_OPENAI_KEY}
    )

vision_client = None
if GOOGLE_CREDENTIALS_JSON and vision:
    try:
        creds_dict = json.loads(GOOGLE_CREDENTIALS_JSON)
        creds = service_account.Credentials.from_service_account_info(creds_dict)
        vision_client = vision.ImageAnnotatorClient(credentials=creds)
    except Exception as e:
        logger.warning("Cloud Vision init failed: %s", e)

def get_drive_service():
    """Google Drive API service (uses same service account JSON as Vision)."""
    if not GOOGLE_CREDENTIALS_JSON or not gdrive_build:
        return None
    try:
        creds_dict = json.loads(GOOGLE_CREDENTIALS_JSON)
        creds = service_account.Credentials.from_service_account_info(
            creds_dict,
            scopes=["https://www.googleapis.com/auth/drive.file"]
        )
        return gdrive_build("drive", "v3", credentials=creds, cache_discovery=False)
    except Exception as e:
        logger.warning("Drive init failed: %s", e)
        return None

def extract_json_array(text: str):
	"""AIのレスポンスからJSON配列を抽出する"""
	if not text:
		raise ValueError("empty response")

	text = text.strip()

	# ```json ... ``` / ``` ... ``` を除去
	text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE)
	text = re.sub(r"\s*```$", "", text)

	# 最初の [ から最後の ] まで抽出
	start = text.find("[")
	end = text.rfind("]")

	if start != -1 and end != -1 and end >= start:
		text = text[start:end + 1]

	data = json.loads(text)

	if not isinstance(data, list):
		data = [data]

	return data

# ==================== PostgreSQL (Supabase) Database ====================
DATABASE_URL = os.getenv("DATABASE_URL")

# グローバルにプールを保持（最小1、最大20コネクション）
db_pool = None
if DATABASE_URL and psycopg2 and pool:
    try:
        # SimpleConnectionPool ではなく ThreadedConnectionPool を使用（スレッドセーフ化）
        db_pool = pool.ThreadedConnectionPool(1, 20, DATABASE_URL, cursor_factory=RealDictCursor)
        logger.info("Database connection pool initialized.")
    except Exception as e:
        logger.error("DB Pool creation failed: %s", e)

def get_db():
    if not DATABASE_URL:
        logger.warning("DATABASE_URL not set. Running in NO-DB mode.")
        return None
    if not psycopg2:
        logger.warning("psycopg2 がインストールされていません。NO-DB モードで動作します。")
        return None
    
    # プールが初期化されていればプールから取得、なければフォールバック
    if db_pool:
        return db_pool.getconn()
    else:
        return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)

def require_db():
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="DATABASE_URL is not configured")
    return conn

from contextlib import contextmanager

@contextmanager
def db_conn():
    """データベース接続を安全に管理するコンテキストマネージャ"""
    conn = require_db()
    try:
        yield conn
    except Exception:
        try:
            conn.rollback()
        except Exception as e:
            logger.warning("DB rollback failed: %s", e)
        raise
    finally:
        release_db(conn)

@contextmanager
def db_transaction():
    """書き込み処理用のトランザクション管理コンテキストマネージャ"""
    conn = require_db()
    try:
        yield conn
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception as e:
            logger.warning("DB rollback failed: %s", e)
        raise
    finally:
        release_db(conn)

def release_db(conn):
    if db_pool:
        db_pool.putconn(conn)
    else:
        conn.close()

def init_db():
    conn = get_db()
    if not conn: return
    with conn.cursor() as cur:
        # Users table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL UNIQUE,
                color TEXT DEFAULT '#c9a961',
                created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
            );
        """)
        # Customers table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS customers (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL UNIQUE,
                code TEXT UNIQUE,
                discount_rate INTEGER NOT NULL DEFAULT 35,
                created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
            );
        """)
        cur.execute("ALTER TABLE customers ADD COLUMN IF NOT EXISTS code TEXT UNIQUE;")
        
        # Invoices table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS invoices (
                id SERIAL PRIMARY KEY,
                invoice_number TEXT NOT NULL UNIQUE,
                customer_name TEXT NOT NULL,
                discount_rate INTEGER NOT NULL,
                total_net_amount INTEGER DEFAULT 0,
                total_tax_amount INTEGER DEFAULT 0,
                total_grand_total INTEGER DEFAULT 0,
                item_count INTEGER DEFAULT 0,
                status TEXT DEFAULT 'draft',
                pdf_storage_path TEXT,
                excel_storage_path TEXT,
                detail_pdf_storage_path TEXT,
                detail_excel_storage_path TEXT,
                created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
                locked_at TIMESTAMP WITH TIME ZONE,
                user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                doc_type TEXT NOT NULL DEFAULT 'delivery'
            );
        """)
        cur.execute("ALTER TABLE invoices ADD COLUMN IF NOT EXISTS user_id INTEGER REFERENCES users(id) ON DELETE SET NULL;")
        cur.execute("ALTER TABLE invoices ADD COLUMN IF NOT EXISTS doc_type TEXT NOT NULL DEFAULT 'delivery';")
        cur.execute("ALTER TABLE invoices ADD COLUMN IF NOT EXISTS customer_code TEXT;")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_invoices_doc_type ON invoices(doc_type);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_invoices_user ON invoices(user_id);")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS invoice_items (
                id SERIAL PRIMARY KEY,
                invoice_id INTEGER NOT NULL REFERENCES invoices(id) ON DELETE CASCADE,
                code TEXT,
                color TEXT,
                size TEXT,
                unit_price INTEGER DEFAULT 0,
                quantity INTEGER DEFAULT 1,
                net_amount INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS api_usage (
                id SERIAL PRIMARY KEY,
                ai_model TEXT NOT NULL,
                image_count INTEGER DEFAULT 0,
                created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS jobs (
                id UUID PRIMARY KEY,
                type TEXT NOT NULL,
                status TEXT NOT NULL,
                payload JSONB NOT NULL,
                result JSONB,
                error TEXT,
                created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
            );
            CREATE INDEX IF NOT EXISTS idx_jobs_status_created ON jobs(status, created_at DESC);
            CREATE INDEX IF NOT EXISTS idx_jobs_created_at ON jobs(created_at);
            CREATE TABLE IF NOT EXISTS invoice_sequences (
                doc_type TEXT NOT NULL,
                yyyymm TEXT NOT NULL,
                current_seq INTEGER NOT NULL DEFAULT 1,
                PRIMARY KEY (doc_type, yyyymm)
            );
        """)
        
        # Initial users if empty
        cur.execute("SELECT COUNT(*) FROM users;")
        if cur.fetchone()['count'] == 0:
            cur.execute("INSERT INTO users (name, color) VALUES (%s, %s), (%s, %s);", 
                       ("緒方", "#c9a961", "高橋", "#6b9bd1"))
        
        cur.execute("SELECT COUNT(*) FROM customers")
        if cur.fetchone()['count'] == 0:
            cur.executemany("INSERT INTO customers (name, discount_rate) VALUES (%s, %s)", [
                ("株式会社 タム 御中", 35),
                ("株式会社 サンプル 御中", 40),
            ])
    try:
        conn.commit()
    finally:
        release_db(conn)

def generate_invoice_number(doc_type='delivery'):
    with db_conn() as conn, conn.cursor() as cur:
        res = generate_invoice_number_safe(cur, doc_type)
        conn.commit()
    return res

def generate_invoice_number_safe(cur, doc_type):
    prefix_code = DOC_TYPE_PREFIXES.get(doc_type, "LJ")
    now = get_jst_now()
    month = now.strftime("%Y%m")
    prefix = f"{prefix_code}-{month}"

    cur.execute("""
        INSERT INTO invoice_sequences (doc_type, yyyymm, current_seq)
        VALUES (%s, %s, 1)
        ON CONFLICT (doc_type, yyyymm)
        DO UPDATE SET current_seq = invoice_sequences.current_seq + 1
        RETURNING current_seq
    """, (doc_type, month))

    seq = cur.fetchone()["current_seq"]
    return f"{prefix}-{seq:04d}"

# ==================== Job Management Helpers ====================
def db_create_job(job_type: str, payload: dict):
    jid = str(uuid.uuid4())
    conn = get_db()
    if not conn: return jid
    import random
    try:
        with conn.cursor() as cur:
            # 50回に1回程度の確率で、7日以上前の古いジョブを掃除する
            if random.random() < 0.02:
                try:
                    cur.execute("DELETE FROM jobs WHERE created_at < NOW() - INTERVAL '7 days'")
                except Exception as e:
                    logger.warning("jobs gc failed: %s", e)

            cur.execute(
                "INSERT INTO jobs (id, type, status, payload) VALUES (%s, %s, %s, %s)",
                (jid, job_type, 'pending', json.dumps(payload))
            )
            conn.commit()
    finally:
        release_db(conn)
    return jid

def db_update_job(jid: str, status: str, result: dict = None, error: str = None):
    conn = get_db()
    if not conn: return
    try:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE jobs SET status=%s, result=%s, error=%s, updated_at=NOW() WHERE id=%s",
                (status, json.dumps(result) if result else None, error, jid)
            )
            conn.commit()
    finally:
        release_db(conn)

def db_get_job(jid: str):
    conn = get_db()
    if not conn: return None
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM jobs WHERE id=%s", (jid,))
            row = cur.fetchone()
    finally:
        release_db(conn)
    if row:
        row = dict(row)
        if isinstance(row['payload'], str): row['payload'] = json.loads(row['payload'])
        if row['result'] and isinstance(row['result'], str): row['result'] = json.loads(row['result'])
    return row

# Initialize DB on startup
try:
    init_db()
except Exception as e:
    logger.error("DB Init Failed: %s", e)

# ==================== Local Font Setup (PDF高速化) ====================
FONT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "static", "fonts"))
_font_css_cache: Optional[str] = None

def _ensure_fonts_downloaded():
    """Noto Sans JP の woff2 を一度だけダウンロードしてローカルに保存。
    PDF 生成のたびに Google Fonts を叩かないための高速化。"""
    os.makedirs(FONT_DIR, exist_ok=True)
    marker = os.path.join(FONT_DIR, ".ready")
    if os.path.exists(marker):
        return True
    try:
        # User-Agent を woff2 対応のものにしてリクエストすると woff2 URL が返る
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"
            )
        }
        css_url = "https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@400;700&display=swap"
        css_resp = requests.get(css_url, headers=headers, timeout=15)
        css_resp.raise_for_status()
        css_text = css_resp.text

        # @font-face ブロックを切り出して、必要な weight (400/700) かつ日本語サブセットだけ拾う
        blocks = re.findall(r"@font-face\s*\{[^}]+\}", css_text)
        wanted_weights = {"400", "700"}
        downloaded = []
        for blk in blocks:
            weight_m = re.search(r"font-weight:\s*(\d+)", blk)
            url_m = re.search(r"url\((https://[^)]+\.woff2)\)", blk)
            if not (weight_m and url_m):
                continue
            weight = weight_m.group(1)
            if weight not in wanted_weights:
                continue
            # 日本語サブセット (Unicode range に CJK Unified Ideographs などを含む) のみ採用。
            # シンプルには U+30...4E... 4F... 等を含むブロックを判定。
            if "U+4E00" not in blk and "U+30" not in blk:
                continue
            fname = f"NotoSansJP-{weight}.woff2"
            fpath = os.path.join(FONT_DIR, fname)
            if not os.path.exists(fpath):
                font_resp = requests.get(url_m.group(1), headers=headers, timeout=30)
                font_resp.raise_for_status()
                with open(fpath, "wb") as f:
                    f.write(font_resp.content)
            downloaded.append((weight, fname))
            wanted_weights.discard(weight)
            if not wanted_weights:
                break

        if downloaded:
            with open(marker, "w") as f:
                f.write(",".join(f"{w}:{n}" for w, n in downloaded))
            return True
    except Exception as e:
        logger.warning("Font download failed (PDF will use fallback): %s", e)
    return False

def get_pdf_font_css() -> str:
    """PDF テンプレート埋め込み用の @font-face CSS を返す。
    ローカルフォントがあればそれを使い、無ければ空文字（システムフォントにフォールバック）。"""
    global _font_css_cache
    if _font_css_cache is not None:
        return _font_css_cache

    _ensure_fonts_downloaded()
    parts = []
    for weight in ("400", "700"):
        fpath = os.path.join(FONT_DIR, f"NotoSansJP-{weight}.woff2")
        if os.path.exists(fpath):
            file_url = "file:///" + fpath.replace(os.sep, "/").lstrip("/")
            parts.append(
                f"@font-face {{ font-family: 'Noto Sans JP'; font-style: normal; "
                f"font-weight: {weight}; src: url('{file_url}') format('woff2'); }}"
            )
    _font_css_cache = "\n".join(parts)
    return _font_css_cache

# 起動時にフォントを取得（失敗してもアプリは起動する）
try:
    _ensure_fonts_downloaded()
except Exception as e:
    logger.warning("Font preload skipped: %s", e)

# ==================== FastAPI App ====================
app = FastAPI()
app.mount("/static", StaticFiles(directory="static"), name="static")
cookie_sec = APIKeyCookie(name="laura_session", auto_error=False)
templates = Jinja2Templates(directory="templates")

# ==================== CSRF Protection (Origin/Referer check) ====================
# 状態変更メソッドに対してリクエスト元オリジンをチェックし、
# Cookie 認証 + SameSite=Lax と組み合わせて CSRF を防ぐ。
_STATE_CHANGING_METHODS = {"POST", "PUT", "DELETE", "PATCH"}
# 環境変数で許可オリジンを上書き可能（カンマ区切り）。未指定時は同一オリジンのみ許可。
_EXTRA_ALLOWED_ORIGINS = {
    o.strip().rstrip("/")
    for o in os.getenv("ALLOWED_ORIGINS", "").split(",")
    if o.strip()
}

@app.middleware("http")
async def origin_check_middleware(request: Request, call_next):
    if request.method in _STATE_CHANGING_METHODS:
        # 自分自身が想定するオリジン（プロキシ経由対応のため Host + scheme から再構築）
        forwarded_proto = request.headers.get("x-forwarded-proto") or request.url.scheme
        host = request.headers.get("host", "")
        same_origin = f"{forwarded_proto}://{host}".rstrip("/") if host else None

        origin = (request.headers.get("origin") or "").rstrip("/")
        referer = request.headers.get("referer") or ""

        allowed = {same_origin} if same_origin else set()
        allowed |= _EXTRA_ALLOWED_ORIGINS

        ok = False
        if origin:
            ok = origin in allowed
        elif referer:
            # Referer は full URL なので origin 部分だけ抽出
            try:
                from urllib.parse import urlparse
                p = urlparse(referer)
                ref_origin = f"{p.scheme}://{p.netloc}".rstrip("/")
                ok = ref_origin in allowed
            except Exception:
                ok = False
        else:
            # ブラウザ以外（ヘルスチェック・curl など）は Origin/Referer が無い場合がある。
            # /login のみ Origin/Referer 無しでも許容（ブラウザによっては送らないため）。
            if request.url.path in {"/login"}:
                ok = True

        if not ok:
            logger.warning("CSRF blocked: method=%s path=%s origin=%r referer=%r",
                           request.method, request.url.path, origin, referer)
            return JSONResponse(status_code=403, content={"detail": "CSRF check failed"})

    return await call_next(request)

_credentials_configured = bool(APP_USERNAME and APP_PASSWORD)
if not _credentials_configured and not DEV_MODE:
    logger.error("APP_USERNAME / APP_PASSWORD が未設定です。本番起動を拒否します。"
                 " ローカル開発で test/test ログインを許可するには DEV_MODE=true を設定してください。")

def _set_session_cookie(resp, token: str):
    resp.set_cookie(
        key="laura_session",
        value=token,
        max_age=60*60*24*365,
        httponly=True,
        secure=True,
        samesite="lax"
    )

def get_session_token():
    if _credentials_configured:
        return hashlib.sha256(f"{APP_USERNAME}:{APP_PASSWORD}:laurajapan".encode()).hexdigest()
    # DEV_MODE時のみ使用される固定トークン
    return hashlib.sha256(b"laura-dev-mode-token").hexdigest()

def authenticate(token: Annotated[str, Depends(cookie_sec)]):
    if not _credentials_configured and not DEV_MODE:
        # 認証情報が未設定で DEV_MODE でもない場合は全リクエストを拒否
        raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="Auth not configured")
    valid_token = get_session_token()
    if not token or not secrets.compare_digest(token, valid_token):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Not authenticated")
    return APP_USERNAME or "dev"

# /login ブルートフォース対策: IP単位で 1分あたり 8 回まで
_login_attempts: Dict[str, List[float]] = {}
_LOGIN_WINDOW_SEC = 60
_LOGIN_MAX_ATTEMPTS = 8

def _check_login_rate_limit(request: Request) -> bool:
    ip = request.client.host if request.client else "unknown"
    now = time.time()
    bucket = _login_attempts.setdefault(ip, [])
    # 古い記録を破棄
    cutoff = now - _LOGIN_WINDOW_SEC
    bucket[:] = [t for t in bucket if t > cutoff]
    if len(bucket) >= _LOGIN_MAX_ATTEMPTS:
        return False
    bucket.append(now)
    # メモリ肥大防止: 一定数を超えたら他IPの古い記録を一斉GC
    if len(_login_attempts) > 1024:
        for k in list(_login_attempts.keys()):
            _login_attempts[k] = [t for t in _login_attempts[k] if t > cutoff]
            if not _login_attempts[k]:
                del _login_attempts[k]
    return True

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    return templates.TemplateResponse(request=request, name="login.html", context={"error": ""})

@app.post("/login", response_class=HTMLResponse)
async def login_post(request: Request, username: str = Form(...), password: str = Form(...)):
    if not _check_login_rate_limit(request):
        return templates.TemplateResponse(
            request=request, name="login.html",
            context={"error": "試行回数が多すぎます。しばらく待ってから再試行してください。"},
            status_code=429,
        )

    if not _credentials_configured:
        if not DEV_MODE:
            return templates.TemplateResponse(request=request, name="login.html", context={"error": "サーバー設定エラー: 認証情報が未設定です"})
        # DEV_MODE: test/test のみ許可
        if username == "test" and password == "test":
            resp = RedirectResponse(url="/", status_code=status.HTTP_302_FOUND)
            _set_session_cookie(resp, get_session_token())
            return resp
        return templates.TemplateResponse(request=request, name="login.html", context={"error": "IDまたはパスワードが間違っています"})

    is_correct_username = secrets.compare_digest(username.encode("utf8"), APP_USERNAME.encode("utf8"))
    is_correct_password = secrets.compare_digest(password.encode("utf8"), APP_PASSWORD.encode("utf8"))

    if is_correct_username and is_correct_password:
        resp = RedirectResponse(url="/", status_code=status.HTTP_302_FOUND)
        _set_session_cookie(resp, get_session_token())
        return resp
    else:
        return templates.TemplateResponse(request=request, name="login.html", context={"error": "IDまたはパスワードが間違っています"})

@app.get("/logout")
async def logout():
    resp = RedirectResponse(url="/login", status_code=status.HTTP_302_FOUND)
    resp.delete_cookie("laura_session")
    return resp

@app.get("/", response_class=HTMLResponse)
async def index(request: Request, token: Annotated[str, Depends(cookie_sec)] = None):
    try:
        authenticate(token)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=status.HTTP_302_FOUND)
    return templates.TemplateResponse(request=request, name="index.html")

@app.get("/manifest.json")
async def get_manifest():
    return FileResponse("manifest.json", media_type="application/manifest+json")

@app.get("/sw.js")
async def get_sw():
    return FileResponse("sw.js", media_type="application/javascript", headers={"Service-Worker-Allowed": "/"})

@app.get("/health")
async def health():
    return {"status": "alive", "time": get_jst_now().isoformat()}

# ==================== Dashboard API ====================
@app.get("/api/dashboard")
def get_dashboard(username: Annotated[str, Depends(authenticate)]):
    with db_conn() as conn, conn.cursor() as cur:
        # DBセッションをJSTに設定
        try:
            cur.execute("SET TIME ZONE 'Asia/Tokyo'")
        except: pass
        
        now = get_jst_now()
        # 今月の開始時刻 (JST)
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        # 前月の開始時刻 (JST)
        last_month_start = (month_start - datetime.timedelta(days=1)).replace(day=1)
        
        # 今月の統計（全伝票タイプ）
        cur.execute(
            """SELECT doc_type, COUNT(*) as cnt, COALESCE(SUM(total_grand_total),0) as total
               FROM invoices WHERE created_at >= %s GROUP BY doc_type""",
            (month_start,)
        )
        this_month_rows = cur.fetchall()
        
        # 前月の統計（全伝票タイプ）
        cur.execute(
            """SELECT doc_type, COUNT(*) as cnt, COALESCE(SUM(total_grand_total),0) as total
               FROM invoices WHERE created_at >= %s AND created_at < %s GROUP BY doc_type""",
            (last_month_start, month_start)
        )
        last_month_rows = cur.fetchall()
        
        # 返品はマイナスとして集計する
        return_types = {'return', 'prov_return'}
        
        def aggregate(rows):
            total_cnt = 0
            total_amt = 0
            delivery_cnt = 0; delivery_amt = 0
            return_cnt = 0; return_amt = 0
            for r in rows:
                dt = r['doc_type']
                cnt = r['cnt']
                amt = abs(r['total']) # 絶対値で扱う
                total_cnt += cnt
                if dt in return_types:
                    return_cnt += cnt
                    return_amt += amt
                    total_amt -= amt  # 返品はマイナス
                else:
                    delivery_cnt += cnt
                    delivery_amt += amt
                    total_amt += amt
            return {
                'invoices': total_cnt,
                'total_amount': total_amt,
                'delivery_count': delivery_cnt,
                'delivery_amount': delivery_amt,
                'return_count': return_cnt,
                'return_amount': return_amt,
            }
        
        this_month = aggregate(this_month_rows)
        last_month = aggregate(last_month_rows)
        
        # 月別推移 (過去12ヶ月, 納品/返品を分離)
        cur.execute(
            """SELECT to_char(created_at, 'YYYY-MM') as month, doc_type,
                      COUNT(*) as cnt, COALESCE(SUM(total_grand_total),0) as total
               FROM invoices GROUP BY month, doc_type ORDER BY month DESC LIMIT 60"""
        )
        monthly_raw = cur.fetchall()
        monthly_map = {}
        for r in monthly_raw:
            m = r['month']
            if m not in monthly_map:
                monthly_map[m] = {'month': m, 'delivery': 0, 'return': 0, 'net': 0, 'cnt': 0}
            amt = abs(r['total']) # 絶対値で扱う
            monthly_map[m]['cnt'] += r['cnt']
            if r['doc_type'] in return_types:
                monthly_map[m]['return'] += amt
                monthly_map[m]['net'] -= amt
            else:
                monthly_map[m]['delivery'] += amt
                monthly_map[m]['net'] += amt
        monthly = sorted(monthly_map.values(), key=lambda x: x['month'], reverse=True)[:12]
        
        # 全期間の取引先別売上トップ5（返品をマイナスで加算）
        cur.execute(
            """SELECT customer_name, doc_type, SUM(total_grand_total) as total
               FROM invoices GROUP BY customer_name, doc_type"""
        )
        cust_raw = cur.fetchall()
        cust_map = {}
        for r in cust_raw:
            name = r['customer_name']
            if name not in cust_map:
                cust_map[name] = 0
            amt = abs(r['total']) # 絶対値で扱う
            if r['doc_type'] in return_types:
                cust_map[name] -= amt
            else:
                cust_map[name] += amt
        top_customers = sorted(
            [{'customer_name': k, 'total': v} for k, v in cust_map.items()],
            key=lambda x: x['total'], reverse=True
        )[:5]
        
        # 全期間の商品別(品番)数量トップ5
        cur.execute(
            """SELECT ii.code, SUM(ii.quantity) as qty
               FROM invoice_items ii JOIN invoices i ON ii.invoice_id = i.id
               GROUP BY ii.code ORDER BY qty DESC LIMIT 5"""
        )
        top_items = cur.fetchall()

        # API利用状況
        cur.execute(
            "SELECT ai_model, COUNT(*) as cnt, COALESCE(SUM(image_count),0) as imgs FROM api_usage WHERE created_at >= %s GROUP BY ai_model",
            (month_start,)
        )
        usage = cur.fetchall()
    
    # コスト計算
    usage_list = []
    for u in usage:
        model = u["ai_model"]
        imgs = u["imgs"]
        if model == "gemini": cost = round(imgs * 0.1, 1)
        elif model in ["azure", "openai"]: cost = round(imgs * 0.5, 1)
        else: cost = 0
        usage_list.append({"model": model, "requests": u["cnt"], "images": imgs, "est_cost_yen": cost})
    
    return {
        "this_month": this_month,
        "last_month": last_month,
        "monthly": monthly,
        "top_customers": [dict(c) if not isinstance(c, dict) else c for c in top_customers],
        "top_items": [dict(i) for i in top_items],
        "api_usage": usage_list
    }

# ==================== User Master API ====================
@app.get("/api/users")
def get_users(username: Annotated[str, Depends(authenticate)]):
    with db_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT id, name, color FROM users ORDER BY id")
        rows = cur.fetchall()
    return [dict(r) for r in rows]

@app.post("/api/users")
def add_user(username: Annotated[str, Depends(authenticate)], name: str = Form(...), color: str = Form("#c9a961")):
    with db_conn() as conn, conn.cursor() as cur:
        cur.execute("INSERT INTO users (name, color) VALUES (%s, %s)", (name, color))
        conn.commit()
    return {"status": "ok"}

@app.delete("/api/users/{uid}")
def delete_user(uid: int, username: Annotated[str, Depends(authenticate)]):
    with db_conn() as conn, conn.cursor() as cur:
        cur.execute("DELETE FROM users WHERE id = %s", (uid,))
        conn.commit()
    return {"status": "ok"}

# ==================== Customer Master API ====================
@app.get("/api/customers")
def get_customers(username: Annotated[str, Depends(authenticate)]):
    with db_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT id, name, code, discount_rate FROM customers ORDER BY id")
        rows = cur.fetchall()
    return [dict(r) for r in rows]

@app.post("/api/customers")
def add_customer(username: Annotated[str, Depends(authenticate)], name: str = Form(...), discount_rate: int = Form(35), code: Optional[str] = Form(None)):
    with db_conn() as conn, conn.cursor() as cur:
        try:
            cur.execute("INSERT INTO customers (name, code, discount_rate) VALUES (%s, %s, %s)", (name, code, discount_rate))
            conn.commit()
        except Exception as e:
            conn.rollback()
            raise HTTPException(400, "登録に失敗しました。店舗コードや店名が既に存在する可能性があります。")
    return {"status": "ok"}

@app.put("/api/customers/{cid}")
def update_customer(cid: int, username: Annotated[str, Depends(authenticate)], name: str = Form(...), discount_rate: int = Form(35), code: Optional[str] = Form(None)):
    with db_conn() as conn, conn.cursor() as cur:
        try:
            cur.execute("UPDATE customers SET name=%s, code=%s, discount_rate=%s WHERE id=%s", (name, code, discount_rate, cid))
            conn.commit()
        except Exception as e:
            conn.rollback()
            raise HTTPException(400, "更新に失敗しました。店舗コードや店名が重複している可能性があります。")
    return {"status": "ok"}

@app.post("/api/customers/import")
def import_customers_api(
    username: Annotated[str, Depends(authenticate)],
    data: str = Form(...),
):
    count = 0
    errors = []

    with db_conn() as conn, conn.cursor() as cur:
        for line_no, line in enumerate(data.strip().split("\n"), start=1):
            parts = line.split("\t")
            if len(parts) < 3:
                errors.append({"line": line_no, "error": "列数不足"})
                continue

            name = parts[0].strip()
            code = parts[1].strip()

            try:
                rate = int(parts[2].strip())
            except ValueError:
                rate = 35

            if not name:
                continue

            try:
                cur.execute("SAVEPOINT sp_import_customer")
                cur.execute("""
                    INSERT INTO customers (name, code, discount_rate)
                    VALUES (%s, %s, %s)
                    ON CONFLICT (name) DO UPDATE SET
                        code = EXCLUDED.code,
                        discount_rate = EXCLUDED.discount_rate
                """, (name, code, rate))
                cur.execute("RELEASE SAVEPOINT sp_import_customer")
                count += 1
            except Exception as e:
                try:
                    cur.execute("ROLLBACK TO SAVEPOINT sp_import_customer")
                except:
                    pass
                errors.append({"line": line_no, "error": str(e)})

        conn.commit()

    return {"status": "ok", "count": count, "errors": errors}

@app.delete("/api/customers/{cid}")
def delete_customer(cid: int, username: Annotated[str, Depends(authenticate)]):
    with db_conn() as conn, conn.cursor() as cur:
        cur.execute("DELETE FROM customers WHERE id = %s", (cid,))
        conn.commit()
    return {"status": "ok"}


# ==================== AI Analysis API ====================
def analyze_images_internal(jid: str, image_parts: list, ai_model: str):
    try:
        db_update_job(jid, 'processing')
        prompt = """
あなたはアパレルブランドのデータ入力アシスタントです。
【重要指示】
1. 画像の前に「--- 画像 X ---」という番号を付与しています。送信された画像の枚数と、出力するJSON配列の要素数は【必ず一致】させてください。全く同じ画像でも省略は厳禁です。
2. 出力データの `source_image_no` には、対象画像の X の数値をそのまま入れてください。
3. 金額は必ず「本体価格（税抜）」を抽出してください。

スキーマ: [{"code": "品番", "color": "カラー", "size": "サイズ", "unit_price": 本体価格の数値, "quantity": 1, "source_image_no": 画像番号}]
""".strip()

        raw_text = ""
        if ai_model == "vision":
            if not vision_client: raise Exception("Cloud VisionのJSONキーが未設定です。")
            items_data = []
            for i, part in enumerate(image_parts):
                image = vision.Image(content=part["data"])
                response = vision_client.text_detection(image=image)
                if response.error.message: raise Exception(f"Vision Error: {response.error.message}")
                raw_text = response.text_annotations[0].description if response.text_annotations else ""
                
                code_match = re.search(r'[A-Za-z0-9]+-[A-Za-z0-9]+', raw_text)
                col_match = re.search(r'(?i)col(?:or)?[\s\.:]*(\d{1,3})', raw_text)
                sz_match = re.search(r'(?i)size[\s\\.:]*([\w]+)', raw_text)
                if not sz_match: sz_match = re.search(r'\b(3[68]|4[02468]|50)\b', raw_text)
                price_match = re.search(r'[¥￥]\s*([\d,]+)', raw_text)
                if not price_match: price_match = re.search(r'\b(\d{1,3}(?:,\d{3})+)\b', raw_text)
                unit_price = 0
                if price_match:
                    try: unit_price = int(price_match.group(1).replace(',', ''))
                    except: pass
                items_data.append({
                    "code": code_match.group(0) if code_match else "-",
                    "color": col_match.group(1) if col_match else "-",
                    "size": sz_match.group(1) if sz_match else "-",
                    "unit_price": unit_price,
                    "quantity": 1,
                    "source_image_no": i,
                })


        elif ai_model in ["azure", "openai"]:
            CHUNK_SIZE = 20
            items_data = []
            
            client = azure_client if ai_model == "azure" else openai_client
            if not client: raise Exception(f"{ai_model.capitalize()} OpenAIのキー/エンドポイントが未設定です。")
            model_name = "gpt-4o" if ai_model == "azure" else "gpt-4o-mini"
            max_tokens_param = {"max_completion_tokens": 4000} if ai_model == "azure" else {"max_tokens": 4000}

            for i in range(0, len(image_parts), CHUNK_SIZE):
                chunk = image_parts[i:i + CHUNK_SIZE]
                content_list = [{"type": "text", "text": prompt}]
                for idx, part in enumerate(chunk):
                    global_idx = i + idx
                    content_list.append({"type": "text", "text": f"--- 画像 {global_idx} ---"})
                    content_list.append({
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:{part['mime_type']};base64,{part['base64']}",
                            "detail": "low"
                        }
                    })
                
                try:
                    response = client.chat.completions.create(
                        model=model_name,
                        messages=[{"role": "user", "content": content_list}],
                        **max_tokens_param
                    )
                    raw_text = response.choices[0].message.content.strip()
                    chunk_data = extract_json_array(raw_text)
                    if not isinstance(chunk_data, list): chunk_data = [chunk_data]
                    items_data.extend(chunk_data)
                except Exception as e:
                    logger.warning("Chunk parsing failed (%s, offset %d): %s", ai_model, i, e)

        elif ai_model == "gemini":
            CHUNK_SIZE = 20
            items_data = []

            for i in range(0, len(image_parts), CHUNK_SIZE):
                chunk = image_parts[i:i + CHUNK_SIZE]
                contents = [prompt]
                for idx, part in enumerate(chunk):
                    global_idx = i + idx
                    contents.append(f"--- 画像 {global_idx} ---")
                    contents.append({"mime_type": part["mime_type"], "data": part["data"]})
                
                try:
                    response = gemini_model.generate_content(contents)
                    raw_text = response.text.strip()
                    chunk_data = extract_json_array(raw_text)
                    if not isinstance(chunk_data, list): chunk_data = [chunk_data]
                    items_data.extend(chunk_data)
                except Exception as e:
                    logger.warning("Chunk parsing failed (gemini, offset %d): %s", i, e)
        
        chunk_data = items_data  # normalize variable name for the rest of the function
        
        # Record usage
        try:
            with db_conn() as conn, conn.cursor() as cur:
                cur.execute("INSERT INTO api_usage (ai_model, image_count) VALUES (%s, %s)", (ai_model, len(image_parts)))
                conn.commit()
        except Exception as e:
            logger.warning("api_usage log failed: %s", e)
        
        db_update_job(jid, 'done', result={"items": chunk_data})
    except Exception as e:
        db_update_job(jid, 'failed', error=str(e))

@app.post("/analyze-images")
async def analyze_images(request: Request, username: Annotated[str, Depends(authenticate)], files: List[UploadFile] = File(...), ai_model: str = Form("gemini")):
    jid = db_create_job('analyze', {"ai_model": ai_model, "file_count": len(files)})
    image_parts = []
    for file in files:
        data = await file.read()
        image_parts.append({"mime_type": file.content_type or "image/jpeg", "data": data, "base64": base64.b64encode(data).decode()})
    # AI 呼び出しは同期的に長時間ブロックするのでイベントループから外す
    await run_in_threadpool(analyze_images_internal, jid, image_parts, ai_model)
    job = db_get_job(jid)
    if job['status'] == 'failed': raise HTTPException(500, job['error'])
    return JSONResponse(job['result'])

@app.post("/api/jobs/analyze")
async def enqueue_analyze(bt: BackgroundTasks, username: Annotated[str, Depends(authenticate)], files: List[UploadFile] = File(...), ai_model: str = Form("gemini")):
    jid = db_create_job('analyze', {"ai_model": ai_model, "file_count": len(files)})
    image_parts = []
    for file in files:
        data = await file.read()
        image_parts.append({"mime_type": file.content_type or "image/jpeg", "data": data, "base64": base64.b64encode(data).decode()})
    bt.add_task(analyze_images_internal, jid, image_parts, ai_model)
    return {"job_id": jid, "status": "pending"}

# ==================== Product Detail Sheet Logic ====================
SIZE_COLUMNS = ["44", "46", "48", "50", "52"]

def apply_a4_print_settings(ws, orientation="portrait", fit_to_width=True, fit_to_height=False):
    """
    Excel出力時の印刷設定をA4に固定する。
    - orientation: "portrait" または "landscape"
    - fit_to_width: 横幅を1ページに収める
    - fit_to_height: 縦も1ページに収める
    """
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = orientation

    ws.sheet_properties.pageSetUpPr.fitToPage = True

    if fit_to_width:
        ws.page_setup.fitToWidth = 1
    else:
        ws.page_setup.fitToWidth = 0

    if fit_to_height:
        ws.page_setup.fitToHeight = 1
    else:
        ws.page_setup.fitToHeight = 0

    # A4に収まりやすい余白
    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.45
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2

    ws.print_options.horizontalCentered = True

def build_detail_excel(invoice_number: str, customer_name: str, items: list, doc_type='delivery') -> bytes:
    """商品明細表 Excel (サイズ別内訳・モダンデザイン)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    titles = DOC_TYPE_TITLES.get(doc_type, DOC_TYPE_TITLES["delivery"])
    ws.title = titles["detail_pdf_title"]
    
    apply_a4_print_settings(ws, orientation="portrait", fit_to_width=True, fit_to_height=False)

    # ===== 共通スタイル定義 =====
    FF = "游ゴシック"
    FONT_TITLE   = Font(name=FF, size=14, bold=True, color="FFFFFF")
    FONT_HEADER  = Font(name=FF, size=9,  bold=True, color="1F2937")
    FONT_BODY    = Font(name=FF, size=10, color="111827")
    FONT_BODY_SM = Font(name=FF, size=9,  color="6B7280")
    FONT_META    = Font(name=FF, size=10, color="374151")
    FONT_INV_NO  = Font(name=FF, size=9,  color="6B7280")

    FILL_TITLE   = PatternFill("solid", fgColor="1F2937")
    FILL_HEADER  = PatternFill("solid", fgColor="F3F4F6")
    FILL_ZEBRA   = PatternFill("solid", fgColor="F9FAFB")

    border_thin = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    border_header_bottom = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='medium', color='1F2937')
    )

    # ===== 列幅 =====
    widths = {'A': 5, 'B': 15, 'C': 7, 'D': 10, 'E': 8, 'F': 5, 'G': 5, 'H': 5, 'I': 5, 'J': 5, 'K': 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    now = get_jst_now()
    reiwa = now.year - 2018

    # ===== タイトルバー (1行目) =====
    ws.merge_cells("A1:K1")
    ws["A1"] = titles["detail"]
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ===== メタ情報 (2-3行目) =====
    ws["A2"] = f"取引先: {customer_name}"
    ws["A2"].font = FONT_META
    ws.merge_cells("A2:E2")

    ws["G2"] = f"伝票番号: {invoice_number}"
    ws["G2"].font = FONT_INV_NO
    ws.merge_cells("G2:K2")
    ws["G2"].alignment = Alignment(horizontal="right", vertical="center")

    ws["I3"] = f"{reiwa}年"
    ws["J3"] = f"{now.month}月"
    ws["K3"] = f"{now.day}日"
    for c in ['I','J','K']:
        ws[f"{c}3"].alignment = Alignment(horizontal="center")
        ws[f"{c}3"].font = FONT_BODY_SM

    # ===== テーブルヘッダー (4行目) =====
    header_row = 4
    ws.row_dimensions[header_row].height = 24
    headers_d = {'A':'No.', 'B':'品番', 'C':'枚数', 'D':'上代', 'E':'カラー', 'F':'44', 'G':'46', 'H':'48', 'I':'50', 'J':'52', 'K':'備考'}
    for col, label in headers_d.items():
        cell = ws[f"{col}{header_row}"]
        cell.value = label
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_header_bottom

    # ===== 商品データ (5行ずつのブロック) =====
    ROWS_PER_BLOCK = 5
    total_blocks = max(1, (len(items) + ROWS_PER_BLOCK - 1) // ROWS_PER_BLOCK)
    
    current_r = header_row + 1
    for b in range(total_blocks):
        for i in range(ROWS_PER_BLOCK):
            r = current_r
            ws.row_dimensions[r].height = 22
            
            item_idx = b * ROWS_PER_BLOCK + i

            # ゼブラ縞
            if item_idx % 2 == 1:
                for col in list('ABCDEFGHIJK'):
                    ws[f"{col}{r}"].fill = FILL_ZEBRA

            # No.（通し番号）
            ws[f"A{r}"] = item_idx + 1
            ws[f"A{r}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"A{r}"].font = FONT_BODY_SM

            if item_idx < len(items):
                item = items[item_idx]
                ws[f"B{r}"] = item.get("code", "")
                ws[f"B{r}"].font = FONT_BODY
                ws[f"C{r}"] = item.get("quantity", 0)
                ws[f"C{r}"].font = FONT_BODY
                ws[f"C{r}"].number_format = '0'
                up = item.get("unit_price", 0) or 0
                if up:
                    ws[f"D{r}"] = up
                    ws[f"D{r}"].number_format = '#,##0'
                ws[f"D{r}"].font = FONT_BODY
                ws[f"E{r}"] = item.get("color", "")
                ws[f"E{r}"].font = FONT_BODY
                
                size_val = str(item.get("size", ""))
                for si, sc in enumerate(SIZE_COLUMNS):
                    if size_val == sc:
                        ws[f"{chr(70+si)}{r}"] = item.get("quantity", 1)
                        ws[f"{chr(70+si)}{r}"].font = FONT_BODY

            for col in list('ABCDEFGHIJK'):
                ws[f"{col}{r}"].border = border_thin
                if col not in ['B', 'K']:  # 品番と備考以外は中央揃え
                    ws[f"{col}{r}"].alignment = Alignment(horizontal="center", vertical="center")
                else:
                    ws[f"{col}{r}"].alignment = Alignment(horizontal="left", vertical="center")
            # 上代は右揃え
            ws[f"D{r}"].alignment = Alignment(horizontal="right", vertical="center", indent=1)
            
            current_r += 1

    # ===== ウィンドウ枠固定 & 印刷設定 =====
    ws.freeze_panes = f"A{header_row + 1}"
    ws.print_title_rows = f'{header_row}:{header_row}'
    last_row = ws.max_row
    ws.print_area = f"A1:K{last_row}"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def build_detail_pdf(invoice_number: str, customer_name: str, items: list, doc_type='delivery') -> bytes:
    """明細 PDF"""
    titles = DOC_TYPE_TITLES.get(doc_type, DOC_TYPE_TITLES["delivery"])
    font_css = get_pdf_font_css()
    rows_html = ""
    for i, item in enumerate(items):
        rows_html += (
            "<tr>"
            f"<td>{i + 1}</td>"
            f"<td>{h(item.get('code'))}</td>"
            f"<td>{h(item.get('color'))}</td>"
            f"<td>{h(item.get('size'))}</td>"
            f"<td>{h(item.get('quantity', 0))}</td>"
            f"<td>¥{int(item.get('unit_price') or 0):,}</td>"
            "</tr>"
        )

    html_str = f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@400;700&display=swap" rel="stylesheet">
<style>{font_css}
body{{font-family:'Noto Sans JP', -apple-system, BlinkMacSystemFont, sans-serif; font-size:12px;}}
.title{{font-size:20px; font-weight:bold; text-align:center; margin-bottom:10px;}}
.meta{{margin-bottom:15px;}}
table{{width:100%; border-collapse:collapse;}}
th,td{{border:1px solid #ccc; padding:6px; text-align:center;}}
th{{background:#f4ecd8;}}
</style>
</head>
<body>
<div class="title">{h(titles['detail'])}</div>
<div class="meta">
  <span>伝票番号: {h(invoice_number)}</span><br>
  <span>取引先: {h(customer_name)}</span>
</div>
<table>
<thead>
<tr><th>No.</th><th>品番</th><th>カラー</th><th>サイズ</th><th>枚数</th><th>上代</th></tr>
</thead>
<tbody>{rows_html}</tbody>
</table>
</body>
</html>"""

    if HTML is None:
        raise RuntimeError("WeasyPrint is not available. PDF generation is disabled.")
    return HTML(string=html_str).write_pdf()

def build_invoice_pdf(invoice_data: dict) -> bytes:
    """納品書 PDF (HTMLテンプレート経由)"""
    from jinja2 import Environment, FileSystemLoader, select_autoescape
    env = Environment(
        loader=FileSystemLoader("templates"),
        autoescape=select_autoescape(["html", "xml"])
    )
    template = env.get_template("invoice_template.html")
    render_data = {**invoice_data, "font_face_css": get_pdf_font_css()}
    html_str = template.render(**render_data)
    if HTML is None:
        raise RuntimeError("WeasyPrint is not available. PDF generation is disabled.")
    return HTML(string=html_str).write_pdf()

def build_invoice_excel(invoice_data: dict, is_preview: bool = False) -> bytes:
    """納品書 Excel (バーコード付き・モダンデザイン)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = invoice_data.get("doc_pdf_title", "伝票")
    
    apply_a4_print_settings(ws, orientation="portrait", fit_to_width=True, fit_to_height=False)

    # ===== 共通スタイル定義 =====
    FF = "游ゴシック"
    FONT_TITLE   = Font(name=FF, size=16, bold=True, color="FFFFFF")
    FONT_HEADER  = Font(name=FF, size=10, bold=True, color="1F2937")
    FONT_LABEL   = Font(name=FF, size=9,  bold=True, color="6B7280")
    FONT_BODY    = Font(name=FF, size=10, color="111827")
    FONT_BODY_SM = Font(name=FF, size=9,  color="6B7280")
    FONT_TOTAL   = Font(name=FF, size=12, bold=True, color="111827")
    FONT_MONEY   = Font(name=FF, size=10, color="111827")

    FILL_TITLE     = PatternFill("solid", fgColor="1F2937")
    FILL_HEADER    = PatternFill("solid", fgColor="F3F4F6")
    FILL_HIGHLIGHT = PatternFill("solid", fgColor="FEF3C7")
    FILL_ZEBRA     = PatternFill("solid", fgColor="F9FAFB")
    FILL_TOTAL     = PatternFill("solid", fgColor="E5E7EB")
    FILL_LABEL_BG  = PatternFill("solid", fgColor="F9FAFB")

    border_thin = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    border_header_bottom = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='medium', color='1F2937')
    )

    # ===== 列幅 (最初に確定させる) =====
    col_widths = {'A': 5, 'B': 14, 'C': 10, 'D': 12, 'E': 26, 'F': 8, 'G': 13, 'H': 15, 'I': 9}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ===== タイトルバー (1行目) =====
    doc_type = invoice_data.get("doc_type", "delivery")
    label_map = {
        "delivery": "納品", "return": "返品",
        "prov_delivery": "仮納品", "prov_return": "仮返品"
    }
    ws.merge_cells("A1:I1")
    ws["A1"] = f"{label_map.get(doc_type, '納品')}伝票"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ===== 伝票番号エリア (2行目) - No.ラベルのみ、値は空欄 =====
    ws["G2"] = "No."
    ws["G2"].font = FONT_LABEL
    ws["G2"].alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells("H2:I2")
    ws["H2"] = ""
    ws["H2"].fill = FILL_HIGHLIGHT
    ws["H2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    # ===== メタ情報 (3-4行目) - 店名を広く =====
    ws.merge_cells("B3:C3")
    ws["B3"] = "コード"
    ws.merge_cells("D3:F3")
    ws["D3"] = "店名"
    ws.merge_cells("G3:I3")
    ws["G3"] = "日付"
    for cell_addr in ["B3", "D3", "G3"]:
        c = ws[cell_addr]
        c.font = FONT_LABEL
        c.fill = FILL_LABEL_BG
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = border_thin

    ws.merge_cells("B4:C4")
    ws["B4"] = invoice_data.get("customer_code", "")
    ws.merge_cells("D4:F4")
    ws["D4"] = invoice_data['customer_name']
    ws.merge_cells("G4:I4")
    ws["G4"] = invoice_data['date']
    for cell_addr in ["B4", "D4", "G4"]:
        c = ws[cell_addr]
        c.font = FONT_BODY
        c.fill = FILL_HIGHLIGHT
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = border_thin

    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 24

    # ===== 店舗バーコード (5行目に独立配置) =====
    ws.row_dimensions[5].height = 38
    store_code = invoice_data.get("customer_code", "")
    if store_code and not is_preview:
        try:
            bc_store = barcode.get('code128', store_code, writer=ImageWriter())
            bc_io_store = BytesIO()
            bc_store.write(bc_io_store, options={
                "write_text": False, "quiet_zone": 2,
                "font_size": 0, "text_distance": 0, "module_height": 10
            })
            img_store = ExcelImage(bc_io_store)
            img_store.width, img_store.height = 160, 36
            marker_store = AnchorMarker(col=1, colOff=pixels_to_EMU(4),
                                        row=4, rowOff=pixels_to_EMU(2))
            img_store.anchor = OneCellAnchor(
                _from=marker_store,
                ext=XDRPositiveSize2D(cx=pixels_to_EMU(160), cy=pixels_to_EMU(36))
            )
            ws.add_image(img_store)
        except Exception as e:
            logger.warning("Store barcode skipped (code=%s): %s", store_code, e)

    # ===== 空白行 (6行目) =====
    ws.row_dimensions[6].height = 6

    # ===== テーブルヘッダー (7行目) =====
    start_row = 7
    headers = ["品番", "カラー", "サイズ", "バーコード", "数量", "単価", "金額", "掛率"]
    h_cols = ["B", "C", "D", "E", "F", "G", "H", "I"]
    ws.row_dimensions[start_row].height = 26
    for col, txt in zip(["A"] + h_cols, ["No."] + headers):
        cell = ws[f"{col}{start_row}"]
        cell.value = txt
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.border = border_header_bottom
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ===== データ行 =====
    dr_val = invoice_data.get('discount_rate') or 0
    rate_label = f"{dr_val}%" if dr_val > 0 else "掛率なし"

    for i, item in enumerate(invoice_data["items"]):
        r = start_row + 1 + i
        ws.row_dimensions[r].height = 36

        if i % 2 == 1:
            for col in ["A","B","C","D","E","F","G","H","I"]:
                ws[f"{col}{r}"].fill = FILL_ZEBRA

        ws[f"A{r}"] = i + 1
        ws[f"A{r}"].font = FONT_BODY_SM
        ws[f"B{r}"] = item["code"]
        ws[f"B{r}"].font = FONT_BODY
        ws[f"C{r}"] = item["color"]
        ws[f"C{r}"].font = FONT_BODY
        ws[f"D{r}"] = item["size"]
        ws[f"D{r}"].font = FONT_BODY
        ws[f"F{r}"] = item["quantity"]
        ws[f"F{r}"].font = FONT_BODY
        ws[f"F{r}"].number_format = '0'
        ws[f"G{r}"] = item["unit_price"]
        ws[f"G{r}"].font = FONT_MONEY
        ws[f"G{r}"].number_format = '#,##0;[Red]△#,##0'
        ws[f"H{r}"] = item["net_amount"]
        ws[f"H{r}"].font = FONT_MONEY
        ws[f"H{r}"].number_format = '#,##0;[Red]△#,##0'
        ws[f"I{r}"] = rate_label
        ws[f"I{r}"].font = FONT_BODY_SM
        
        for col in ["A","B","C","D","E","F","G","H","I"]:
            ws[f"{col}{r}"].border = border_thin

        for col in ["A", "B", "C", "D", "F", "I"]:
            ws[f"{col}{r}"].alignment = Alignment(horizontal="center", vertical="center")
        for col in ["G", "H"]:
            ws[f"{col}{r}"].alignment = Alignment(horizontal="right", vertical="center", indent=1)
        
        if not is_preview:
            try:
                bc_str = f"{item['code']}{item['color']}{item['size']}".replace("-", "")
                ean = barcode.get('code128', bc_str, writer=ImageWriter())
                bc_io = BytesIO()
                ean.write(bc_io, options={
                    "write_text": False, "quiet_zone": 2,
                    "font_size": 0, "text_distance": 0, "module_height": 10
                })
                img = ExcelImage(bc_io)
                img.width, img.height = 160, 36
                marker = AnchorMarker(col=4, colOff=pixels_to_EMU(8),
                                      row=r-1, rowOff=pixels_to_EMU(3))
                img.anchor = OneCellAnchor(
                    _from=marker,
                    ext=XDRPositiveSize2D(cx=pixels_to_EMU(160), cy=pixels_to_EMU(36))
                )
                ws.add_image(img)
            except Exception as e:
                logger.warning("Item barcode skipped (row %d, code=%s): %s", r, item.get("code"), e)

    # ===== 合計エリア =====
    last_r = start_row + len(invoice_data["items"]) + 2
    top_thick = Side(style='medium', color='1F2937')
    for i, (label, val) in enumerate([
        ("小計", invoice_data["total_net_amount"]),
        ("消費税", invoice_data["total_tax_amount"]),
        ("合計金額", invoice_data["total_grand_total"]),
    ]):
        rr = last_r + i
        c_label = ws.cell(row=rr, column=7, value=label)
        c_val = ws.cell(row=rr, column=8, value=val)
        c_val.number_format = '¥#,##0;[Red]△¥#,##0'

        if i == 2:
            c_label.font = FONT_TOTAL
            c_val.font = FONT_TOTAL
            c_label.fill = FILL_TOTAL
            c_val.fill = FILL_TOTAL
            c_label.border = Border(top=top_thick, bottom=top_thick,
                                    left=Side(style='thin', color='D1D5DB'),
                                    right=Side(style='thin', color='D1D5DB'))
            c_val.border = Border(top=top_thick, bottom=top_thick,
                                  left=Side(style='thin', color='D1D5DB'),
                                  right=Side(style='thin', color='D1D5DB'))
        else:
            c_label.font = FONT_BODY
            c_val.font = FONT_MONEY
            c_label.border = border_thin
            c_val.border = border_thin

        c_label.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c_val.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # ===== ウィンドウ枠固定 & 印刷設定 =====
    ws.freeze_panes = f"A{start_row + 1}"
    ws.print_title_rows = f'{start_row}:{start_row}'
    last_row = ws.max_row
    ws.print_area = f"A1:I{last_row}"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()



def build_all_files(invoice_data: dict, is_preview: bool = False) -> dict:
    """4ファイルまとめて生成"""
    doc_type = invoice_data.get("doc_type", "delivery")
    return {
        "pdf": build_invoice_pdf(invoice_data),
        "excel": build_invoice_excel(invoice_data, is_preview=is_preview),
        "detail_pdf": build_detail_pdf(invoice_data["invoice_number"], invoice_data["customer_name"], invoice_data["items"], doc_type),
        "detail_excel": build_detail_excel(invoice_data["invoice_number"], invoice_data["customer_name"], invoice_data["items"], doc_type),
    }

def build_single_file(kind: str, invoice_data: dict) -> bytes:
    """
    指定された1種類のファイルだけを生成する。
    kind:
      - pdf
      - excel
      - detail_pdf
      - detail_excel
    """
    doc_type = invoice_data.get("doc_type", "delivery")

    if kind == "pdf":
        return build_invoice_pdf(invoice_data)

    if kind == "excel":
        return build_invoice_excel(invoice_data)

    if kind == "detail_pdf":
        return build_detail_pdf(
            invoice_data["invoice_number"],
            invoice_data["customer_name"],
            invoice_data["items"],
            doc_type,
        )

    if kind == "detail_excel":
        return build_detail_excel(
            invoice_data["invoice_number"],
            invoice_data["customer_name"],
            invoice_data["items"],
            doc_type,
        )

    raise ValueError(f"Unsupported file kind: {kind}")

def assemble_invoice_data(inv_info: dict, items_input: list, discount_rate: int, doc_type='delivery') -> dict:
    """生成用データ準備"""
    processed = []
    total_net = total_tax = total_grand = 0
    # 掛け率が0（手動入力等）の場合は、掛け率なし（100%）として計算
    rate = (discount_rate / 100.0) if discount_rate > 0 else 1.0
    
    # 追加: 返品系なら数量と金額をマイナスにするための係数
    sign = -1 if doc_type in ['return', 'prov_return'] else 1

    for it in items_input:
        up = it.get("unit_price", 0)
        if isinstance(up, str): up = int(up.replace(',','').replace('¥','').strip() or '0')
        
        # 変更: 絶対値にしてから sign（1 または -1）を掛ける
        qty = abs(int(it.get("quantity") or 1)) or 1
        qty = qty * sign
        
        # Pythonの int() は0方向に切り捨てるため、マイナスでも正しく計算される
        net = int(up * rate * qty)
        tax = int(net * 0.1)
        grand = net + tax
        
        processed.append({
            "code": it.get("code") or "-", "color": it.get("color") or "-", "size": it.get("size") or "-",
            "unit_price": up, "quantity": qty, "net_amount": net, "tax_amount": tax, "grand_total": grand
        })
        total_net += net; total_tax += tax; total_grand += grand
    
    dt = to_jst(inv_info.get("locked_at") or inv_info.get("created_at"))

    titles = DOC_TYPE_TITLES.get(doc_type, DOC_TYPE_TITLES["delivery"])

    return {
        "invoice_number": inv_info["invoice_number"], "customer_name": inv_info["customer_name"], "customer_code": inv_info.get("customer_code", ""),
        "discount_rate": discount_rate, "items": processed,
        "total_net_amount": total_net, "total_tax_amount": total_tax, "total_grand_total": total_grand,
        "issuer": "株式会社 ラウラジャパン", "date": dt.strftime("%Y年%m月%d日"),
        "doc_type": doc_type,
        "doc_title_main": titles["main"],
        "doc_title_detail": titles["detail"],
        "doc_pdf_title": titles["pdf_title"],
        "doc_detail_pdf_title": titles["detail_pdf_title"]
    }

# ==================== Document Generation API ====================
class DocumentRequest(BaseModel):
    invoice_id: Optional[int] = None
    customer_code: Optional[str] = None
    customer_name: str
    discount_rate: int
    items: List[Dict[str, Any]]
    doc_type: Optional[str] = "delivery"
    user_id: Optional[int] = None

@app.post("/api/preview")
def preview_documents(username: Annotated[str, Depends(authenticate)], payload: DocumentRequest):
    """保存せず、HTMLテンプレートをそのまま返す（爆速プレビュー）"""
    inv_data = assemble_invoice_data(
        {"invoice_number": "PREVIEW-" + get_jst_now().strftime("%H%M%S"), "customer_name": payload.customer_name},
        payload.items, payload.discount_rate, payload.doc_type or "delivery"
    )
    
    # WeasyPrintでPDF化せず、Jinja2のHTML文字列をそのまま返す
    from jinja2 import Environment, FileSystemLoader, select_autoescape
    env = Environment(
        loader=FileSystemLoader("templates"),
        autoescape=select_autoescape(["html", "xml"])
    )
    template = env.get_template("invoice_template.html")
    # preview用にはフォントCSSを空にする（ブラウザがGoogle Fontsを読むため）
    render_data = {**inv_data, "font_face_css": ""}
    html_str = template.render(**render_data)
    
    return {
        "preview_only": True,
        "invoice_number": inv_data["invoice_number"],
        "doc_type": inv_data["doc_type"],
        "customer_name": inv_data["customer_name"],
        "customer_code": inv_data.get("customer_code"),
        "discount_rate": inv_data["discount_rate"],
        "items": inv_data["items"],
        "html_preview": html_str,
    }


def _save_invoice_record(payload: "DocumentRequest"):
    """伝票レコードを INSERT/UPDATE して (inv_id, invoice_number, inv_data) を返す共通処理"""
    doc_type = payload.doc_type or "delivery"
    with db_conn() as conn, conn.cursor() as cur:
        if payload.invoice_id:
            cur.execute("SELECT invoice_number, status FROM invoices WHERE id = %s", (payload.invoice_id,))
            row = cur.fetchone()
            if not row: raise Exception("Invoice not found")
            if row["status"] == "locked": raise Exception("確定済みの伝票は編集できません。")
            invoice_number = row["invoice_number"]
        else:
            invoice_number = generate_invoice_number_safe(cur, doc_type)

        inv_data = assemble_invoice_data(
            {"invoice_number": invoice_number, "customer_name": payload.customer_name, "customer_code": payload.customer_code},
            payload.items, payload.discount_rate, doc_type,
        )

        if payload.invoice_id:
            cur.execute("""
                UPDATE invoices SET customer_name=%s, customer_code=%s, discount_rate=%s, total_net_amount=%s, total_tax_amount=%s, total_grand_total=%s,
                item_count=%s, status='draft', locked_at=NULL, doc_type=%s, user_id=%s,
                pdf_storage_path=NULL, excel_storage_path=NULL, detail_pdf_storage_path=NULL, detail_excel_storage_path=NULL
                WHERE id=%s
            """, (payload.customer_name, payload.customer_code, payload.discount_rate, inv_data["total_net_amount"], inv_data["total_tax_amount"], inv_data["total_grand_total"], len(inv_data["items"]), doc_type, payload.user_id, payload.invoice_id))
            cur.execute("DELETE FROM invoice_items WHERE invoice_id=%s", (payload.invoice_id,))
            inv_id = payload.invoice_id
        else:
            cur.execute("""
                INSERT INTO invoices (invoice_number, customer_name, customer_code, discount_rate, total_net_amount, total_tax_amount, total_grand_total, item_count, status, doc_type, user_id)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'draft',%s,%s) RETURNING id
            """, (invoice_number, payload.customer_name, payload.customer_code, payload.discount_rate, inv_data["total_net_amount"], inv_data["total_tax_amount"], inv_data["total_grand_total"], len(inv_data["items"]), doc_type, payload.user_id))
            inv_id = cur.fetchone()['id']

        for item in inv_data["items"]:
            cur.execute("INSERT INTO invoice_items (invoice_id, code, color, size, unit_price, quantity, net_amount) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                         (inv_id, item["code"], item["color"], item["size"], item["unit_price"], item["quantity"], item["net_amount"]))
        conn.commit()
    return inv_id, invoice_number, inv_data


def generate_documents_internal(jid: str, payload_dict: dict):
    try:
        payload = DocumentRequest(**payload_dict)
        doc_type = payload.doc_type or "delivery"
        inv_id, invoice_number, _ = _save_invoice_record(payload)
        result = {
            "invoice_id": inv_id, "invoice_number": invoice_number, "status": "draft", "doc_type": doc_type,
        }
        db_update_job(jid, 'done', result=result)
    except Exception as e:
        db_update_job(jid, 'failed', error=str(e))

@app.post("/generate-documents")
def generate_documents(request: Request, username: Annotated[str, Depends(authenticate)], payload: DocumentRequest):
    """同期APIだが二重生成を避けるため _save_invoice_record で得た inv_data をそのまま使う"""
    jid = db_create_job('generate', payload.model_dump())
    try:
        inv_id, invoice_number, inv_data = _save_invoice_record(payload)
        result = {
            "invoice_id": inv_id, "invoice_number": invoice_number, "status": "draft",
            "doc_type": payload.doc_type or "delivery",
        }
        db_update_job(jid, 'done', result=result)

        return JSONResponse({
            **result,
            "pdf_url": f"/api/history/{inv_id}/pdf",
            "excel_url": f"/api/history/{inv_id}/excel",
            "detail_pdf_url": f"/api/history/{inv_id}/detail-pdf",
            "detail_excel_url": f"/api/history/{inv_id}/detail-excel",
            "pdf_preview_url": f"/api/history/{inv_id}/pdf-preview",
            "detail_pdf_preview_url": f"/api/history/{inv_id}/detail-pdf-preview",
        })
    except Exception as e:
        db_update_job(jid, 'failed', error=str(e))
        raise HTTPException(500, str(e))

@app.post("/api/jobs/generate")
async def enqueue_generate(bt: BackgroundTasks, username: Annotated[str, Depends(authenticate)], payload: DocumentRequest):
    jid = db_create_job('generate', payload.model_dump())
    bt.add_task(generate_documents_internal, jid, payload.model_dump())
    return {"job_id": jid, "status": "pending"}



def lock_invoice_internal(jid: str, inv_id: int, bt: Optional[BackgroundTasks] = None):
    try:
        with db_conn() as conn, conn.cursor() as cur:
            # status が 'locked' でない場合のみ更新を実行（レースコンディション対策）
            cur.execute("""
                UPDATE invoices SET status='locked', locked_at=NOW()
                WHERE id=%s AND status != 'locked'
                RETURNING id
            """, (inv_id,))
            
            if not cur.fetchone():
                db_update_job(jid, 'done', result={"status": "already_locked"})
                return
            conn.commit()

        db_update_job(jid, 'done', result={"status": "locked"})

    except Exception as e:
        db_update_job(jid, 'failed', error=str(e))

@app.post("/api/history/{inv_id}/lock")
async def lock_invoice(inv_id: int, bt: BackgroundTasks, username: Annotated[str, Depends(authenticate)]):
    jid = db_create_job('lock', {"inv_id": inv_id})
    # ロック処理は threadpool で同期実行し、Drive 同期はレスポンス後の BackgroundTasks に乗せる
    await run_in_threadpool(lock_invoice_internal, jid, inv_id, bt)
    job = db_get_job(jid)
    if job['status'] == 'failed': raise HTTPException(500, job['error'])
    return job['result']

@app.post("/api/jobs/lock/{inv_id}")
async def enqueue_lock(bt: BackgroundTasks, inv_id: int, username: Annotated[str, Depends(authenticate)]):
    jid = db_create_job('lock', {"inv_id": inv_id})
    bt.add_task(lock_invoice_internal, jid, inv_id, bt)
    return {"job_id": jid, "status": "pending"}


@app.post("/api/history/{inv_id}/unlock")
async def unlock_invoice(inv_id: int, username: Annotated[str, Depends(authenticate)]):
    with db_conn() as conn, conn.cursor() as cur:
        cur.execute("UPDATE invoices SET status='draft', locked_at=NULL WHERE id=%s", (inv_id,))
        conn.commit()
    return {"status": "draft"}

# ==================== History & Serving API ====================
@app.get("/api/history")
def get_history(username: Annotated[str, Depends(authenticate)], doc_type: Optional[str] = None, user_id: Optional[int] = None):
    with db_conn() as conn, conn.cursor() as cur:
        query = """
            SELECT i.id, i.invoice_number, i.customer_name, i.total_grand_total, i.item_count, i.status, i.doc_type, i.user_id,
            u.name as user_name, u.color as user_color,
            to_char(i.locked_at, 'YYYY-MM-DD"T"HH24:MI:SS') as locked_at,
            to_char(i.created_at, 'YYYY-MM-DD\"T\"HH24:MI:SS') as created_at
            FROM invoices i
            LEFT JOIN users u ON i.user_id = u.id
            WHERE 1=1
        """
        params = []
        if doc_type and doc_type != 'all':
            query += " AND i.doc_type = %s"
            params.append(doc_type)
        if user_id:
            query += " AND i.user_id = %s"
            params.append(user_id)

        query += " ORDER BY i.id DESC"
        cur.execute(query, tuple(params))
        rows = cur.fetchall()
    return [dict(r) for r in rows]

# kind: "pdf" | "excel" | "detail_pdf" | "detail_excel"
KIND_CONFIG = {
    "pdf":          {"path_field": "pdf_storage_path",          "mime": "application/pdf",                                                            "ext": "pdf",  "files_key": "pdf"},
    "excel":        {"path_field": "excel_storage_path",        "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",         "ext": "xlsx", "files_key": "excel"},
    "detail_pdf":   {"path_field": "detail_pdf_storage_path",   "mime": "application/pdf",                                                            "ext": "pdf",  "files_key": "detail_pdf"},
    "detail_excel": {"path_field": "detail_excel_storage_path", "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",         "ext": "xlsx", "files_key": "detail_excel"},
}

def serve_file(inv_id: int, kind: str, disposition: str = "attachment"):
    cfg = KIND_CONFIG[kind]
    with db_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT * FROM invoices WHERE id=%s", (inv_id,))
        inv = cur.fetchone()
        if not inv:
            raise HTTPException(404, "Not found")
        cur.execute("SELECT code,color,size,unit_price,quantity FROM invoice_items WHERE invoice_id=%s", (inv_id,))
        items = [dict(r) for r in cur.fetchall()]

    customer = safe_filename(inv.get("customer_name") or "Unknown")

    dt = to_jst(inv.get("locked_at") or inv.get("created_at"))
    date_str = f"{dt.month}月{dt.day}日"

    doc_type = inv.get("doc_type", "delivery")
    title_map = DOC_TYPE_TITLES.get(doc_type, DOC_TYPE_TITLES["delivery"])
    if kind in ("pdf", "excel"):
        label = title_map["pdf_title"]
    else:
        label = title_map["detail_pdf_title"]

    fname = f'{customer}_{date_str}_{label}.{cfg["ext"]}'
    encoded_fname = urllib.parse.quote(fname)

    # inline / attachment を切り替え
    content_disposition = f"{disposition}; filename*=UTF-8''{encoded_fname}"

    # キャッシュがあればそれを使う
    if inv.get(cfg["path_field"]):
        try:
            content = storage_download(inv[cfg["path_field"]])
            return Response(
                content=content,
                media_type=cfg["mime"],
                headers={
                    "Content-Disposition": content_disposition,
                    "Cache-Control": "private, max-age=300",
                }
            )
        except Exception as e:
            logger.warning("Storage download failed, regenerating: %s", e)

    # キャッシュがない、または失敗した場合はオンデマンドで再生成
    dr = inv.get("discount_rate") or 100
    inv_data = assemble_invoice_data(dict(inv), items, dr, inv.get("doc_type", "delivery"))
    content = build_single_file(cfg["files_key"], inv_data)

    return Response(
        content=content,
        media_type=cfg["mime"],
        headers={
            "Content-Disposition": content_disposition,
            "Cache-Control": "private, max-age=60",
        }
    )


@app.get("/api/history/{inv_id}/pdf")
def dl_pdf(inv_id: int, username: Annotated[str, Depends(authenticate)]):
    return serve_file(inv_id, "pdf")

@app.get("/api/history/{inv_id}/excel")
def dl_excel(inv_id: int, username: Annotated[str, Depends(authenticate)]):
    return serve_file(inv_id, "excel")

@app.get("/api/history/{inv_id}/detail-pdf")
def dl_detail_pdf(inv_id: int, username: Annotated[str, Depends(authenticate)]):
    return serve_file(inv_id, "detail_pdf")

@app.get("/api/history/{inv_id}/detail-excel")
def dl_detail_excel(inv_id: int, username: Annotated[str, Depends(authenticate)]):
    return serve_file(inv_id, "detail_excel")

@app.get("/api/history/{inv_id}/pdf-preview", response_class=HTMLResponse)
def preview_pdf(inv_id: int, username: Annotated[str, Depends(authenticate)]):
    """保存済みの伝票のPDFプレビューの代わりに、HTMLテンプレートをレンダリングして返す"""
    with db_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT * FROM invoices WHERE id=%s", (inv_id,))
        inv = cur.fetchone()
        if not inv:
            raise HTTPException(404, "Not found")
        cur.execute("SELECT code, color, size, unit_price, quantity FROM invoice_items WHERE invoice_id=%s", (inv_id,))
        items = [dict(r) for r in cur.fetchall()]

    dr = inv.get("discount_rate") or 100
    inv_data = assemble_invoice_data(dict(inv), items, dr, inv.get("doc_type", "delivery"))

    from jinja2 import Environment, FileSystemLoader, select_autoescape
    env = Environment(loader=FileSystemLoader("templates"), autoescape=select_autoescape(["html", "xml"]))
    template = env.get_template("invoice_template.html")
    
    # プレビュー表示用なのでWebフォント(Google Fonts)を読み込ませるため font_face_css は空にする
    render_data = {**inv_data, "font_face_css": ""}
    return template.render(**render_data)

@app.get("/api/history/{inv_id}/detail-pdf-preview", response_class=HTMLResponse)
def preview_detail_pdf(inv_id: int, username: Annotated[str, Depends(authenticate)]):
    """保存済みの明細表PDFプレビューの代わりに、動的にHTMLを生成して返す"""
    with db_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT * FROM invoices WHERE id=%s", (inv_id,))
        inv = cur.fetchone()
        if not inv:
            raise HTTPException(404, "Not found")
        cur.execute("SELECT code, color, size, unit_price, quantity FROM invoice_items WHERE invoice_id=%s", (inv_id,))
        items = [dict(r) for r in cur.fetchall()]

    doc_type = inv.get("doc_type", "delivery")
    titles = DOC_TYPE_TITLES.get(doc_type, DOC_TYPE_TITLES["delivery"])
    
    rows_html = ""
    for i, item in enumerate(items):
        rows_html += (
            "<tr>"
            f"<td>{i + 1}</td>"
            f"<td>{h(item.get('code'))}</td>"
            f"<td>{h(item.get('color'))}</td>"
            f"<td>{h(item.get('size'))}</td>"
            f"<td>{h(item.get('quantity', 0))}</td>"
            f"<td>¥{int(item.get('unit_price') or 0):,}</td>"
            "</tr>"
        )

    html_str = f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@400;700&display=swap" rel="stylesheet">
<style>
body{{font-family:'Noto Sans JP', -apple-system, BlinkMacSystemFont, sans-serif; font-size:12px;}}
.title{{font-size:20px; font-weight:bold; text-align:center; margin-bottom:10px;}}
.meta{{margin-bottom:15px;}}
table{{width:100%; border-collapse:collapse;}}
th,td{{border:1px solid #ccc; padding:6px; text-align:center;}}
th{{background:#f4ecd8;}}
</style>
</head>
<body>
<div class="title">{h(titles['detail'])}</div>
<div class="meta">
  <span>伝票番号: {h(inv['invoice_number'])}</span><br>
  <span>取引先: {h(inv['customer_name'])}</span>
</div>
<table>
<thead>
<tr><th>No.</th><th>品番</th><th>カラー</th><th>サイズ</th><th>枚数</th><th>上代</th></tr>
</thead>
<tbody>{rows_html}</tbody>
</table>
</body>
</html>"""

    return html_str

def get_doc_type_folder_label(doc_type: str) -> str:
    """Drive保存用のフォルダ名ラベルを返す"""
    return {
        "delivery": "納品",
        "return": "返品",
        "prov_delivery": "仮納品",
        "prov_return": "仮返品",
    }.get(doc_type, "納品")

def safe_filename(name: str) -> str:
    """ファイル名に使えない文字を置換する"""
    name = name or "unknown"
    return re.sub(r'[\\/:*?"<>|]+', "_", name).strip()

def get_invoice_generated_files(inv_id: int):
    """伝票の生成ファイル（4種類）を取得する。確定済みでキャッシュがあればそれを優先する"""
    with db_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT * FROM invoices WHERE id=%s", (inv_id,))
        inv = cur.fetchone()
        if not inv:
            raise HTTPException(404, "Invoice not found")
        cur.execute("SELECT code, color, size, unit_price, quantity FROM invoice_items WHERE invoice_id=%s", (inv_id,))
        items = [dict(r) for r in cur.fetchall()]

    inv = dict(inv)
    path_map = {
        "pdf": "pdf_storage_path",
        "excel": "excel_storage_path",
        "detail_pdf": "detail_pdf_storage_path",
        "detail_excel": "detail_excel_storage_path",
    }

    # 確定済みかつ4ファイルのキャッシュが揃っている場合はStorageを優先
    if inv.get("status") == "locked" and all(inv.get(col) for col in path_map.values()):
        try:
            files = {
                key: storage_download(inv[col])
                for key, col in path_map.items()
            }
            return inv, files
        except Exception as e:
            logger.warning("Cached files unavailable, regenerating: %s", e)

    # キャッシュなし・下書き・取得失敗時は再生成
    dr = inv.get("discount_rate") or 100
    inv_data = assemble_invoice_data(inv, items, dr, inv.get("doc_type", "delivery"))
    files = build_all_files(inv_data)
    return inv, files

@app.get("/api/history/{inv_id}/zip")
def dl_zip(inv_id: int, username: Annotated[str, Depends(authenticate)]):
    """伝票に関連する4ファイルをZIPにまとめてダウンロードする"""
    try:
        inv, files = get_invoice_generated_files(inv_id)
        titles = DOC_TYPE_TITLES.get(inv.get("doc_type", "delivery"), DOC_TYPE_TITLES["delivery"])
        
        inv_num = safe_filename(inv.get("invoice_number") or "Unknown")
        customer = safe_filename(inv.get("customer_name") or "Unknown")
        
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as z:
            z.writestr(f"{inv_num}_{customer}_{titles['pdf_title']}.pdf", files["pdf"])
            z.writestr(f"{inv_num}_{customer}_{titles['pdf_title']}.xlsx", files["excel"])
            z.writestr(f"{inv_num}_{customer}_{titles['detail_pdf_title']}.pdf", files["detail_pdf"])
            z.writestr(f"{inv_num}_{customer}_{titles['detail_pdf_title']}.xlsx", files["detail_excel"])
        
        zip_buffer.seek(0)
        
        # === ここからZIPファイル名の変更 ===
        dt = to_jst(inv.get("locked_at") or inv.get("created_at"))
        date_str = f"{dt.month}月{dt.day}日"
        
        zip_name = f"{customer}_{date_str}.zip"
        # === 変更ここまで ===

        # 日本語ファイル名を安全にヘッダーに含めるためのエンコード (RFC 6266)
        encoded_filename = urllib.parse.quote(zip_name)
        
        return Response(
            content=zip_buffer.getvalue(),
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )
    except Exception as e:
        logger.exception("ZIP download failed for inv_id %s: %s", inv_id, e)
        raise HTTPException(status_code=500, detail=f"ZIP生成に失敗しました: {str(e)}")

@app.get("/api/history/{inv_id}/items")
def get_history_items(inv_id: int, username: Annotated[str, Depends(authenticate)]):
    with db_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT code, color, size, unit_price, quantity FROM invoice_items WHERE invoice_id = %s", (inv_id,))
        rows = cur.fetchall()
        cur.execute("SELECT invoice_number, customer_name, discount_rate, doc_type FROM invoices WHERE id = %s", (inv_id,))
        inv = cur.fetchone()
    if not inv: raise HTTPException(status_code=404, detail="Invoice not found")
    return {"invoice_number": inv["invoice_number"], "customer_name": inv["customer_name"], "discount_rate": inv["discount_rate"], "doc_type": inv.get("doc_type", "delivery"), "items": [dict(r) for r in rows]}

@app.delete("/api/history/{inv_id}")
def delete_history(inv_id: int, username: Annotated[str, Depends(authenticate)]):
    with db_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT pdf_storage_path, excel_storage_path, detail_pdf_storage_path, detail_excel_storage_path FROM invoices WHERE id = %s", (inv_id,))
        row = cur.fetchone()
        if not row: raise HTTPException(404, "Invoice not found")
        
        # 確定済みであっても削除を許可する（403回避）
        
        if STORAGE_BUCKET and supabase_client:
            for col in ['pdf_storage_path', 'excel_storage_path', 'detail_pdf_storage_path', 'detail_excel_storage_path']:
                path = row.get(col)
                if not path:
                    continue
                try:
                    supabase_client.storage.from_(STORAGE_BUCKET).remove([path])
                except Exception as e:
                    logger.warning("Failed cleanup for %s (%s): %s", inv_id, col, e)

        cur.execute("DELETE FROM invoice_items WHERE invoice_id = %s", (inv_id,))
        cur.execute("DELETE FROM invoices WHERE id = %s", (inv_id,))
        conn.commit()
    return {"status": "ok"}

def upload_drive_internal(jid: str, inv_id: int):
    try:
        db_update_job(jid, 'processing')
        if not GDRIVE_WEBHOOK_URL: raise Exception("GDRIVE_WEBHOOK_URLが未設定です")
        
        # 常に最新を生成してドライブにアップロード
        inv, files = get_invoice_generated_files(inv_id)

        doc_type = inv.get("doc_type", "delivery")
        dt = to_jst(inv.get("locked_at") or inv.get("created_at"))

        folder_path = [
            str(dt.year),
            dt.strftime("%Y-%m"),
            get_doc_type_folder_label(doc_type)
        ]

        inv_num = safe_filename(inv.get("invoice_number") or "Unknown")
        cust = safe_filename(inv.get("customer_name") or "無名")
        
        title_map = DOC_TYPE_TITLES.get(inv.get("doc_type", "delivery"), DOC_TYPE_TITLES["delivery"])
        main_label   = title_map["pdf_title"]
        detail_label = title_map["detail_pdf_title"]

        uploaded = []
        targets = [
            ("pdf",          f"{inv_num}_{cust}_{main_label}.pdf",   "application/pdf", files["pdf"]),
            ("excel",        f"{inv_num}_{cust}_{main_label}.xlsx",  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", files["excel"]),
            ("detail_pdf",   f"{inv_num}_{cust}_{detail_label}.pdf", "application/pdf", files["detail_pdf"]),
            ("detail_excel", f"{inv_num}_{cust}_{detail_label}.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", files["detail_excel"]),
        ]
        for typ, fn, mime, data in targets:
            resp = requests.post(GDRIVE_WEBHOOK_URL, json={
                "rootFolderId": GDRIVE_FOLDER_ID, 
                "folderPath": folder_path,
                "filename": fn, 
                "mime": mime, 
                "base64": base64.b64encode(data).decode()
            }, timeout=60)
            if resp.status_code != 200: raise Exception(f"GAS Error {resp.status_code}")
            uploaded.append({"type": typ, "name": fn, "url": resp.json().get("url")})
        db_update_job(jid, 'done', result={"uploaded": uploaded})
    except Exception as e:
        db_update_job(jid, 'failed', error=str(e))

@app.post("/api/history/{inv_id}/upload-drive")
def upload_to_drive(inv_id: int, username: Annotated[str, Depends(authenticate)]):
    jid = db_create_job('drive_upload', {"invoice_id": inv_id})
    upload_drive_internal(jid, inv_id)
    job = db_get_job(jid)
    if job['status'] == 'failed': return JSONResponse(status_code=500, content={"detail": job['error']})
    return job['result']

@app.post("/api/jobs/drive-upload")
def enqueue_drive_upload(bt: BackgroundTasks, username: Annotated[str, Depends(authenticate)], inv_id: int = Form(...)):
    jid = db_create_job('drive_upload', {"invoice_id": inv_id})
    bt.add_task(upload_drive_internal, jid, inv_id)
    return {"job_id": jid, "status": "pending"}

@app.get("/api/jobs")
def get_jobs(username: Annotated[str, Depends(authenticate)], limit: int = 20):
    with db_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT id, type, status, created_at, updated_at FROM jobs ORDER BY created_at DESC LIMIT %s", (limit,))
        rows = [dict(r) for r in cur.fetchall()]
    return rows

@app.get("/api/jobs/{job_id}")
def get_job_status(job_id: str, username: Annotated[str, Depends(authenticate)]):
    job = db_get_job(job_id)
    if not job: raise HTTPException(404, "Job not found")
    return job

@app.get("/api/history/{inv_id}/meta")
def get_history_meta(inv_id: int, username: Annotated[str, Depends(authenticate)]):
    """軽量な履歴詳細（メタデータのみ）取得API"""
    with db_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT * FROM invoices WHERE id=%s", (inv_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Not found")
        inv = dict(row)
        cur.execute("SELECT code, color, size, unit_price, quantity FROM invoice_items WHERE invoice_id=%s", (inv_id,))
        items = [dict(r) for r in cur.fetchall()]

    def iso_or_none(v):
        if not v: return None
        if hasattr(v, "isoformat"): return v.isoformat()
        return str(v)

    return {
        "invoice_id": inv["id"],
        "invoice_number": inv["invoice_number"],
        "customer_name": inv["customer_name"],
        "customer_code": inv.get("customer_code"),
        "discount_rate": inv["discount_rate"],
        "status": inv["status"],
        "doc_type": inv.get("doc_type", "delivery"),
        "total_net_amount": inv.get("total_net_amount", 0),
        "total_tax_amount": inv.get("total_tax_amount", 0),
        "total_grand_total": inv.get("total_grand_total", 0),
        "item_count": inv.get("item_count", len(items)),
        "created_at": iso_or_none(inv.get("created_at")),
        "locked_at": iso_or_none(inv.get("locked_at")),
        "items": items,
        # ダウンロード用URL
        "pdf_url": f"/api/history/{inv_id}/pdf",
        "excel_url": f"/api/history/{inv_id}/excel",
        "detail_pdf_url": f"/api/history/{inv_id}/detail-pdf",
        "detail_excel_url": f"/api/history/{inv_id}/detail-excel",
        # iframe表示専用URL（Content-Disposition: inline）
        "pdf_preview_url": f"/api/history/{inv_id}/pdf-preview",
        "detail_pdf_preview_url": f"/api/history/{inv_id}/detail-pdf-preview",
    }

@app.get("/api/history/{inv_id}/data")
def get_history_data(inv_id: int, username: Annotated[str, Depends(authenticate)]):
    """(互換性維持) 全ドキュメントをBase64で含む詳細API"""
    with db_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT * FROM invoices WHERE id=%s", (inv_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Not found")
        inv = dict(row)
        cur.execute("SELECT code,color,size,unit_price,quantity FROM invoice_items WHERE invoice_id=%s", (inv_id,))
        items = [dict(r) for r in cur.fetchall()]

    inv_data = assemble_invoice_data(inv, items, inv["discount_rate"], inv.get("doc_type", "delivery"))
    files = build_all_files(inv_data)

    def iso_or_none(v):
        if not v: return None
        if hasattr(v, "isoformat"): return v.isoformat()
        return str(v)

    return {
        "invoice_id": inv["id"],
        "invoice_number": inv["invoice_number"],
        "customer_name": inv["customer_name"],
        "customer_code": inv.get("customer_code"),
        "discount_rate": inv["discount_rate"],
        "status": inv["status"],
        "doc_type": inv.get("doc_type", "delivery"),
        "total_net_amount": inv.get("total_net_amount", 0),
        "total_tax_amount": inv.get("total_tax_amount", 0),
        "total_grand_total": inv.get("total_grand_total", 0),
        "item_count": inv.get("item_count", len(items)),
        "created_at": iso_or_none(inv.get("created_at")),
        "locked_at": iso_or_none(inv.get("locked_at")),
        "items": items,
        "pdf_base64": base64.b64encode(files["pdf"]).decode(),
        "excel_base64": base64.b64encode(files["excel"]).decode(),
        "detail_pdf_base64": base64.b64encode(files["detail_pdf"]).decode(),
        "detail_excel_base64": base64.b64encode(files["detail_excel"]).decode(),
    }


@app.get("/manual", response_class=HTMLResponse)
def get_manual(request: Request):
    return templates.TemplateResponse(request=request, name="manual.html")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
